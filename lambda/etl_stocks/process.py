import re
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

_NO_RATE_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

logger = logging.getLogger(__name__)

# Matches: <number> <unit>, e.g. "1000 ML", "1 KG", "8 GMS", "100 GM", "1 LTR"
# Alternation is ordered longest-first to avoid partial matches (LTR > LT > L, GMS > GM).
_UNIT_RE = re.compile(
    r'(\d+(?:\.\d+)?)\s*(GMS|GM|KG|ML|LTR|LT|L)\b',
    re.IGNORECASE,
)

_OUTPUT_HEADERS = [
    'Brand', 'Technical', 'Packing Size', 'Packing Configuration',
    'Available Nos', 'Conversion Factor', 'Available Cases', 'Available Qty',
    'Branch', 'Special Packing Mention', 'Entry Date',
    'Rate', 'Stock Valuation',
]

_PURCHASE_PL = 'Purchase Price List'


def _load_rates(rates_path: str) -> dict[str, float]:
    """Return {product_string: rate} from the Product Master, Purchase Price List only."""
    wb = openpyxl.load_workbook(rates_path, data_only=True)
    ws = wb.active
    rates: dict[str, float] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        prod, brand, pl, rate = row[1], row[3], row[4], row[5]
        if brand and rate is not None and pl == _PURCHASE_PL and prod:
            rates[str(prod).strip()] = float(rate)
    logger.info("Loaded %d product rates from %s", len(rates), rates_path)
    return rates


def _to_int_if_whole(value: float) -> int | float:
    return int(value) if value % 1 == 0 else value


def _parse_product(product: str, brand: str) -> tuple[str, float, str, str]:
    """
    Returns (technical, packing_size, packing_config, packing_spec).

    packing_size is always in base unit (grams or ml).
    packing_config is 'Grams' or 'ml'.
    packing_spec is the text remaining after the size token (e.g. 'TIN', 'BOX', 'NA').

    Handles three layouts:
      1. Standard: Technical - Brand - Size [- Spec]
      2. Embedded: Technical - Brand Size Spec   (IMIX pattern)
      3. Multi-part technical: T1 - T2 - Brand - Size  (VIVAYA PLUS pattern)
    """
    parts = [p.strip() for p in product.split(' - ')]
    brand_upper = brand.upper()

    brand_idx = None
    after_text = ''

    for i, part in enumerate(parts):
        part_upper = part.upper()
        if part_upper == brand_upper:
            brand_idx = i
            # Join remaining parts; strip each to handle stray spaces
            after_text = ' '.join(p.strip() for p in parts[i + 1:])
            break
        # Brand embedded at start of segment with size/spec appended (e.g. "IMIX 8 GMS TIN")
        if part_upper.startswith(brand_upper + ' '):
            brand_idx = i
            after_text = part[len(brand):].strip()
            break

    technical = ' - '.join(parts[:brand_idx]) if brand_idx is not None else product

    if not after_text:
        return technical, 0.0, '', 'NA'

    m = _UNIT_RE.search(after_text)
    if not m:
        return technical, 0.0, '', 'NA'

    size_num = float(m.group(1))
    unit = m.group(2).upper()
    spec = after_text[m.end():].strip() or 'NA'

    if unit == 'KG':
        return technical, size_num * 1000, 'gms', spec
    if unit in ('GMS', 'GM'):
        return technical, size_num, 'gms', spec
    if unit in ('LTR', 'LT', 'L'):
        return technical, size_num * 1000, 'ml', spec
    # ML
    return technical, size_num, 'ml', spec


def process_stock_file(src_path: str, dst_path: str, entry_date: datetime = None, rates_path: str = None) -> int:
    """
    Read a Current Stock Balances xlsx and write a Stock-Processed xlsx.

    Rows with the same (Brand, Technical, Packing Size, Packing Configuration,
    Branch, Special Packing Mention) are merged: Available Nos is summed and
    Available Cases / Available Qty are recalculated from the total.

    Skips rows where Brand column is empty (packaging materials, labels, etc.).
    Returns the number of data rows written.
    """
    if entry_date is None:
        entry_date = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    rates = _load_rates(rates_path) if rates_path else {}

    src_wb = openpyxl.load_workbook(src_path, data_only=True)
    src_ws = src_wb.active

    # merged[(brand, technical, packing_size, packing_config, branch, spec)]
    #   = {'nos': total, 'cf': first_cf, 'product': raw_product_string}
    merged: dict[tuple, dict] = {}

    for row in src_ws.iter_rows(min_row=6, values_only=True):
        branch = row[0]
        product = row[1]
        brand = row[3]
        qty = row[4]
        cf = row[5]

        if not brand or not str(brand).strip():
            continue
        if not product or not str(product).strip():
            continue

        brand = str(brand).strip()
        product = str(product).strip()

        technical, packing_size, packing_config, packing_spec = _parse_product(product, brand)
        branch_str = str(branch).strip() if branch else ''
        available_nos = qty or 0
        conversion_factor = cf or 0

        key = (brand, technical, packing_size, packing_config, branch_str, packing_spec)

        if key in merged:
            merged[key]['nos'] += available_nos
        else:
            merged[key] = {'nos': available_nos, 'cf': conversion_factor, 'product': product}

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = 'Sheet1'
    dst_ws.append(_OUTPUT_HEADERS)

    rows = []
    for key, agg in merged.items():
        brand, technical, packing_size, packing_config, branch_str, packing_spec = key
        total_nos = agg['nos']
        cf = agg['cf']
        available_cases = (total_nos / cf) if cf else 0
        available_qty = packing_size * total_nos
        rate = rates.get(agg['product'])
        stock_valuation = _to_int_if_whole(total_nos * rate) if rate is not None else None

        rows.append({
            'brand': brand,
            'technical': technical,
            'packing_size': packing_size,
            'packing_configuration': packing_config,
            'available_nos': total_nos,
            'conversion_factor': cf,
            'available_cases': available_cases,
            'available_qty': available_qty,
            'branch': branch_str,
            'special_packing_mention': packing_spec,
            'entry_date': entry_date,
            'rate': rate,
            'stock_valuation': stock_valuation,
        })

        dst_ws.append([
            brand,
            technical,
            _to_int_if_whole(packing_size),
            packing_config,
            total_nos,
            cf,
            _to_int_if_whole(available_cases),
            _to_int_if_whole(available_qty),
            branch_str,
            packing_spec,
            entry_date,
            rate,
            stock_valuation,
        ])
        if rate is None:
            for cell in dst_ws[dst_ws.max_row]:
                cell.fill = _NO_RATE_FILL

    dst_wb.save(dst_path)
    logger.info("Wrote %d rows to %s (merged from source)", len(rows), dst_path)
    return rows
