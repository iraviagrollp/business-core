import json
import logging
import os
import tempfile
from urllib.parse import unquote_plus

import boto3
import openpyxl
import psycopg2

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')
secrets = boto3.client('secretsmanager')

_BUCKET = os.environ['DATA_BUCKET']
_RAW_PREFIX = os.environ.get('RAW_PREFIX', 'raw/')
_PROCESSED_PREFIX = os.environ.get('PROCESSED_PREFIX', 'processed/')
_FILE_PREFIX = 'Customer Accounts Export File'

# Column indices (0-based), row 1 is header, data starts row 2
# Delivery Address sheet: [0]=Name, [3]=DLAddress3, [4]=DLCity, [5]=DLState, [7]=DLPIN, [9]=DLMobileNo
# General sheet: [0]=Name, [2]=Code (party code, e.g. "ANK001")

_STATE_MAP = {
    '37-Andhra Pradesh': 'AP',
    '36-Telangana': 'TG',
}


def _get_db_conn():
    secret = json.loads(
        secrets.get_secret_value(SecretId=os.environ['DB_SECRET_ARN'])['SecretString']
    )
    return psycopg2.connect(
        host=secret['host'],
        port=secret.get('port', 5432),
        dbname=secret['dbname'],
        user=secret['username'],
        password=secret['password'],
    )


def lambda_handler(event, context):
    for record in event.get('Records', []):
        bucket = record['s3']['bucket']['name']
        key = unquote_plus(record['s3']['object']['key'])
        logger.info('S3 event: s3://%s/%s', bucket, key)

        filename = key.split('/')[-1]
        if not (filename.startswith(_FILE_PREFIX) and filename.endswith('.xlsx')):
            logger.info('Skipping: %s', key)
            continue

        _process(bucket, key, filename)


def _process(bucket: str, key: str, filename: str):
    archive_key = _PROCESSED_PREFIX + 'raw/' + filename

    with tempfile.TemporaryDirectory() as tmp:
        src_path = os.path.join(tmp, filename)
        logger.info('Downloading s3://%s/%s', bucket, key)
        s3.download_file(bucket, key, src_path)
        rows = _parse(src_path)
        logger.info('Parsed %d customer rows', len(rows))

    conn = _get_db_conn()
    try:
        _upsert(conn, rows)
        conn.commit()
        logger.info('Upserted %d rows into customer_details', len(rows))
    finally:
        conn.close()

    s3.copy_object(Bucket=bucket, CopySource={'Bucket': bucket, 'Key': key}, Key=archive_key)
    s3.delete_object(Bucket=bucket, Key=key)
    logger.info('Archived source to s3://%s/%s', bucket, archive_key)


def _build_code_lookup(wb) -> dict:
    """Return uppercase-name -> party code dict from the 'General' sheet.

    General sheet layout (0-indexed, header row 1, data from row 2):
      [0]=Name, [2]=Code (party code, e.g. "ANK001")

    Names are uppercased to match the normalization applied to customer_name
    in _parse(). Code is cast to str and stripped; blank/None -> None (NULL).
    """
    ws = wb['General']
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or '').strip().upper()
        if not name:
            continue
        code_raw = row[2]
        if code_raw is None:
            code = None
        else:
            code = str(code_raw).strip() or None
        lookup[name] = code
    logger.info('Built code lookup: %d entries (%d with a code)',
                len(lookup), sum(1 for c in lookup.values() if c is not None))
    return lookup


def _build_delivery_lookup(wb) -> dict:
    """Return uppercase-name -> address-fields dict from the 'Delivery Address' sheet.

    Delivery Address sheet layout (0-indexed, header row 1, data from row 2):
      [0]=Name, [3]=DLAddress3 (district), [4]=DLCity, [5]=DLState,
      [7]=DLPIN, [9]=DLMobileNo

    Names are uppercased to match the normalization applied elsewhere.
    All address transforms are identical to those previously applied in _parse():
      - district/city: title-cased, blank -> None
      - state: mapped via _STATE_MAP, missing key -> None
      - pin: str-stripped, blank -> None
      - mobile_no: int/float -> str(int()), str -> strip spaces; last 10 digits if > 10

    First occurrence of a name wins when the same name appears more than once.
    """
    ws = wb.active
    lookup = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or '').strip()
        if not name:
            continue
        upper_name = name.upper()
        if upper_name in lookup:
            # first occurrence wins
            continue

        district = str(row[3] or '').strip()
        city = str(row[4] or '').strip()
        state_raw = str(row[5] or '').strip()
        pin = str(row[7]).strip() if row[7] is not None else None
        mobile_raw = row[9]
        if mobile_raw is None:
            mobile_no = None
        elif isinstance(mobile_raw, (int, float)):
            mobile_no = str(int(mobile_raw))
        else:
            mobile_no = str(mobile_raw).replace(' ', '') or None
        if mobile_no and len(mobile_no) > 10:
            mobile_no = mobile_no[-10:]

        lookup[upper_name] = {
            'district': district.title() if district else None,
            'city': city.title() if city else None,
            'state': _STATE_MAP.get(state_raw),
            'pin': pin or None,
            'mobile_no': mobile_no,
        }

    logger.info('Built delivery lookup: %d entries', len(lookup))
    return lookup


def _parse(src_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(src_path, data_only=True)

    # Build party-code lookup from the General sheet — this is the authoritative
    # customer list: every account in General gets a row WITH its code.
    code_lookup = _build_code_lookup(wb)

    # Build address lookup from the Delivery Address sheet (wb.active).
    # Used for enrichment only; customers absent from here still get inserted.
    delivery_lookup = _build_delivery_lookup(wb)

    logger.info(
        'Customer counts — General: %d, Delivery Address: %d',
        len(code_lookup), len(delivery_lookup),
    )

    # Union of both sets: no customer is lost regardless of which sheet they
    # appear on.  General-only customers get address fields = None.
    # Delivery-only customers get customer_code = None.
    all_names = set(code_lookup) | set(delivery_lookup)
    logger.info('Total unique customers after union: %d', len(all_names))

    rows = []
    for name in sorted(all_names):
        addr = delivery_lookup.get(name, {})
        rows.append({
            'customer_name': name,
            'district': addr.get('district'),
            'city': addr.get('city'),
            'state': addr.get('state'),
            'pin': addr.get('pin'),
            'mobile_no': addr.get('mobile_no'),
            'customer_code': code_lookup.get(name),
        })

    return rows


def _upsert(conn, rows: list[dict]):
    with conn.cursor() as cur:
        for row in rows:
            cur.execute(
                """
                INSERT INTO customer_details
                    (customer_name, district, city, state, pin, mobile_no, customer_code)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (customer_name) DO UPDATE SET
                    district      = EXCLUDED.district,
                    city          = EXCLUDED.city,
                    state         = EXCLUDED.state,
                    pin           = EXCLUDED.pin,
                    mobile_no     = EXCLUDED.mobile_no,
                    customer_code = EXCLUDED.customer_code,
                    updated_at    = NOW()
                """,
                (row['customer_name'], row['district'], row['city'],
                 row['state'], row['pin'], row['mobile_no'], row['customer_code']),
            )
